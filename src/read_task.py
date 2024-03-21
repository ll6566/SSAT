# -*- encoding=utf8 -*-
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

from config.config import root_path, logging
from openpyxl.reader.excel import load_workbook


def eRead_get_cell_coordinate_test(value_to_find):
    """
    获取测试用例表值坐标
    """
    value_to_find = str(value_to_find).replace(" ", "").replace("*", "").replace("\n", "")
    # 加载工作簿
    workbook = load_workbook(root_path + 'TestList.xlsx')
    # 获取工作表
    sheet = workbook['测试用例']

    # 遍历工作表中的所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            # 如果找到了值，返回它的坐标
            if cell.value == value_to_find:
                return cell.column, cell.row


def e_input_x_y_get_value_test(x, y):
    """
    获取坐标值
    """
    # 加载工作簿
    workbook = load_workbook(root_path + 'TestList.xlsx')
    # 获取工作表
    sheet = workbook['测试用例']
    data = sheet.cell(row=y, column=x).value
    return data


def eRead_get_cell_coordinate(value_to_find):
    """
    获取任务表值坐标
    """
    value_to_find = str(value_to_find).replace(" ","").replace("*","").replace("\n","")
    # 加载工作簿
    workbook = load_workbook(root_path + 'TestList.xlsx')
    # 获取工作表
    sheet = workbook['任务表']

    # 遍历工作表中的所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            # 如果找到了值，返回它的坐标
            if cell.value == value_to_find:
                return cell.column, cell.row


def e_input_x_y_get_value(x, y):
    """
    获取坐标值
    """
    # 加载工作簿
    workbook = load_workbook(root_path + 'TestList.xlsx')
    # 获取工作表
    sheet = workbook['任务表']
    data = sheet.cell(row=y, column=x).value
    return data


def eRead_monitor_time():
    """
    读取Excel表中监控启动间隔时间
    return 间隔时间
    """
    x, y = eRead_get_cell_coordinate("监控")
    read = e_input_x_y_get_value(x, y + 1)
    x, y = eRead_get_cell_coordinate("监控检查间隔时间")
    time = e_input_x_y_get_value(x, y + 1)
    if read == "不启动":
        return False, False,""
    elif read == "单独启动监控":
        return True, False, time
    return True, True, time


def eRead_sub_com():
    """
    读取Excel表中的usb控制器COM端口
    return usbHub端口1,usbHub端口2
    """
    x, y = eRead_get_cell_coordinate("hub控制端口1")
    hub1 = str(e_input_x_y_get_value(x, y + 1)).replace(" ","").replace("\n","")

    x, y = eRead_get_cell_coordinate("hub控制端口2")
    hub2 = str(e_input_x_y_get_value(x, y + 1)).replace(" ","").replace("\n","")

    # if str(hub1) is None or str(hub1).count("COM") < 1 or str(hub2) is None or str(hub2).count("COM") < 1:
    #     logging.error("usbHub控制器端口不存在！！")
    #     exit()
    return hub1, hub2


def eRead_wifi_ssid_password():
    """
    读取Excel表中的WIFI SSID 密码
    return SSID,密码
    """
    x, y = eRead_get_cell_coordinate("测试设备连接WifiSSID")
    ssid = e_input_x_y_get_value(x, y + 1)

    x, y = eRead_get_cell_coordinate("测试设备连接WIFI密码")
    password = e_input_x_y_get_value(x, y + 1)
    return ssid, password


def eRead_software_link():
    """
    读取Excel表中的 软件链接
    return 软件链接
    """
    x, y = eRead_get_cell_coordinate("测试软件链接")
    read = e_input_x_y_get_value(x + 1, y)
    if read is None or "https" not in read or ".7z" not in read:
        logging.warning("软件地址有误!!")
        exit()
    return str(read).replace(" ", "").replace("\n", "").replace("\t", "")


def eRead_test_task_dormant():
    """
    读取Excel 休眠设备任务信息
    return 设备任务信息
    """
    task_objet = []
    index = 0
    while True:
        index += 1
        indexX, indexY = eRead_get_cell_coordinate("启动")
        RunType = e_input_x_y_get_value(indexX, indexY + index)
        if RunType is None: break
        if RunType == "工作": continue

        indexX, indexY = eRead_get_cell_coordinate("排位编号")
        Ids = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("机器物理编号")
        Sign = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("Device")
        Device = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("使用的正式登录账号")
        Account = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("连接hub编号")
        HubNumber = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("测试前置条件")
        Preconditions = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("自动化执行节点")
        Nodes = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("启动测试项目")
        BeginName = e_input_x_y_get_value(indexX, indexY + index)

        infos = {
            'Ids': Ids,
            'Sign': Sign,
            'Device': Device,
            'Account': Account,
            'HubNumber': HubNumber,
            'Preconditions': Preconditions,
            'Nodes': Nodes,
            'BeginName': BeginName
        }
        task_objet.append(infos)
    return task_objet


def eRead_test_task():
    """
    读取Excel 设备任务信息
    return 设备任务信息
    """
    task_objet = []
    index = 0
    while True:
        index += 1
        indexX, indexY = eRead_get_cell_coordinate("启动")
        RunType = e_input_x_y_get_value(indexX, indexY + index)
        if RunType is None: break
        if RunType == "休眠": continue
        indexX, indexY = eRead_get_cell_coordinate("排位编号")
        Ids = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("机器物理编号")
        Sign = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("Device")
        Device = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("使用的正式登录账号")
        Account = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("连接hub编号")
        HubNumber = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("测试前置条件")
        Preconditions = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("自动化执行节点")
        Nodes = e_input_x_y_get_value(indexX, indexY + index)
        indexX, indexY = eRead_get_cell_coordinate("启动测试项目")
        BeginName = e_input_x_y_get_value(indexX, indexY + index)

        infos = {
            'Ids': Ids,
            'Sign': Sign,
            'Device': Device,
            'Account': Account,
            'HubNumber': HubNumber,
            'Preconditions': Preconditions,
            'Nodes': Nodes,
            'BeginName': BeginName
        }
        task_objet.append(infos)
    return task_objet


def eRead_dev_type():
    """
    获取设备的状态信息
    return 在线设备 离线设备
    """
    devList = eRead_test_task()
    res = os.popen("adb devices").read()
    tDevList = []
    fDevList = []
    for devs in devList:
        if devs['Device'] is None or devs['Device'] == "":
            logging.error("============================")
            logging.error("设备device为空！！")
            exit()
        if devs['Device'] in res:
            Device = "【{}】".format("在线") + devs['Device']
            devOb = {
                '排位编号': devs['Ids'],
                '设备device': Device,
                '物理机位置': devs['Sign'],
                '测试项目': devs['BeginName'],
                'hub端口': devs['HubNumber'],
                '自动化步骤': devs['Nodes'],
            }
            tDevList.append(devOb)
        else:
            Device = "【{}】".format("离线") + devs['Device']
            devOb = {
                '排位编号': devs['Ids'],
                '设备device': Device,
                '物理机位置': devs['Sign'],
                '测试项目': devs['BeginName'],
                'hub端口': devs['HubNumber'],
                '自动化步骤': devs['Nodes'],
            }
            fDevList.append(devOb)

    return tDevList, fDevList


def eRead_user_notes():
    """
    获取用户名称、测试备注
    """
    x, y = eRead_get_cell_coordinate("测试用户")
    testUser = e_input_x_y_get_value(x, y + 1)

    x, y = eRead_get_cell_coordinate("备注信息(自定义标题)")
    testNotes = e_input_x_y_get_value(x, y + 1)

    return testUser, testNotes


def eRead_test_activity():
    indexX, indexY = eRead_get_cell_coordinate_test("Activity")
    valueList = []
    for xyIndex in range(9999):
        try:
            act = e_input_x_y_get_value_test(indexX, indexY + (xyIndex + 1))
            name = e_input_x_y_get_value_test(indexX + 1, indexY + (xyIndex + 1))
            if act == "-": continue
            if act is None: break
            value = {name: act}
            valueList.append(value)
        except Exception as e:
            logging.error(e)
            break
    return valueList
